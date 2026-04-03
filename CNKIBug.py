"""
CNKI_Bug_dev - 中国知网论文标题爬虫
版本: 0.1.5
作者: Kaffu_Alcaid
打包说明: pyinstaller --onefile --console --name CNKIBug CNKIBug.py
"""

import sys
import os
import shutil
import subprocess
import winreg


def _popup_error(lines: list[str]):
    echo_cmds = []
    for ln in lines:
        if ln.strip():
            echo_cmds.append(f"echo {ln}")
        else:
            echo_cmds.append("echo.")

    inner = " & ".join(echo_cmds) + " & echo. & pause "
    subprocess.Popen(
        ["cmd.exe", "/k", f"color 4E & {inner}"],
        creationflags=subprocess.CREATE_NEW_CONSOLE,
    )


try:
    from playwright.sync_api import (
        sync_playwright,
        TimeoutError as PlaywrightTimeoutError,
        Error as PlaywrightError,
    )
    import openpyxl
    from rich.console import Console
    from rich.progress import (
        Progress,
        SpinnerColumn,
        BarColumn,
        TextColumn,
        TimeElapsedColumn,
        MofNCompleteColumn,
    )
except ImportError as _err:
    if sys.platform == "win32":
        _popup_error([
            "==============================================",
            " [致命错误] 程序核心组件加载失败！",
            "----------------------------------------------",
            f" 缺失模块: {_err}",
            "",
            " 可能原因：您运行的不是完整的 exe 文件，",
            " 或 exe 文件已损坏。",
            "",
            " 解决方法：",
            "   请重新下载完整的 CNKIBug.exe 文件，",
            "   不要解压、不要移动内部文件，直接双击运行。",
            "==============================================",
        ])
    else:
        print(f"[FATAL] 缺少依赖: {_err}")
        print("请运行: pip install playwright openpyxl rich && playwright install chromium")
    sys.exit(1)

import time
import random
import re

_console = Console(highlight=False)

# v0.1.6: 标志位——_scrape_keyword 在 Progress 循环内截获 Ctrl+C 后置 True，
# 通知 scrape_cnki 当前关键词已有半成品数据入账，应停止并保存
_stop_requested = False

_EDGE_PATHS = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\Application\msedge.exe"),
]


def _edge_installed() -> bool:
    if any(os.path.isfile(p) for p in _EDGE_PATHS):
        return True
    return shutil.which("msedge") is not None


def get_real_desktop_path() -> str:
    if sys.platform != "win32":
        return os.path.join(os.path.expanduser("~"), "Desktop")
    try:
        key = winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders",
        )
        val, _ = winreg.QueryValueEx(key, "Desktop")
        winreg.CloseKey(key)
        return os.path.expandvars(val)
    except Exception:
        return os.path.join(os.path.expanduser("~"), "Desktop")


def check_env():
    if sys.platform != "win32":
        playwright_path = os.path.join(
            os.path.expanduser("~"), "AppData", "Local", "ms-playwright"
        )
        if not os.path.exists(playwright_path):
            _console.print("\n[yellow][环境缺失] 请先在终端运行: playwright install chromium[/yellow]\n")
            sys.exit(0)
        return

    if not _edge_installed():
        _popup_error([
            "==============================================",
            " [环境缺失] 未检测到 Microsoft Edge 浏览器！",
            "----------------------------------------------",
            " 本程序需要使用 Microsoft Edge 来抓取网页数据。",
            " Windows 10/11 通常已预装，若您已卸载请重新安装。",
            "",
            " 请用浏览器打开以下地址，下载并安装 Edge：",
            "",
            "   https://www.microsoft.com/zh-cn/edge/download",
            "",
            " 安装完成后，关闭此窗口，重新双击程序即可！",
            "==============================================",
        ])
        sys.exit(0)


def _check_captcha(page, progress: Progress | None = None) -> bool:
    captcha_selectors = [
        "div.captcha",
        "div#captcha",
        "iframe[src*='captcha']",
        "div.slide-verify",
        "div.passcode-area",
    ]
    # 用 progress.console.print 输出，不调用 stop()/start()，避免计时器归零
    out = progress.console.print if progress else _console.print
    for sel in captcha_selectors:
        if page.query_selector(sel):
            out("\n" + "!" * 50)
            out("  [bold yellow][!] 检测到人机验证！请切换到浏览器窗口处理：[/bold yellow]")
            out("  · 如果是【滑块验证】：按住滑块向右拖动到底")
            out("  · 如果是【图片验证码】：按提示点击或输入字符")
            out("  · 完成验证后，回到此窗口按 [回车键] 继续...")
            out("!" * 50)
            input()
            return True
    return False


def _sanitize_name(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|\[\]]', '_', text)


def _scrape_keyword(page, keyword: str, max_pages: int) -> list:
    global _stop_requested
    results = []
    _console.print(f"\n[bold][*][/bold] 目标关键词：[bold cyan]{keyword}[/bold cyan]")

    # v0.1.5: 导航步骤拆分为独立 status 块，移除 KeyboardInterrupt 捕获使其向上传递
    try:
        with _console.status(
            "[bold magenta]少女祈祷中...[/bold magenta]",
            spinner="bouncingBar",
        ):
            page.goto("https://www.cnki.net/", timeout=30000)
            page.wait_for_load_state("domcontentloaded", timeout=20000)
    except PlaywrightTimeoutError:
        _console.print("[yellow][!] 预热请求超时，跳过该关键词。[/yellow]")
        return results
    except PlaywrightError as e:
        _console.print(f"[yellow][!] 预热请求失败: {e}，跳过该关键词。[/yellow]")
        return results
    _check_captcha(page)

    try:
        with _console.status(
            "[bold magenta]少女祈祷中...[/bold magenta]",
            spinner="bouncingBar",
        ):
            page.goto("https://kns.cnki.net/kns8s/", timeout=30000)
            page.wait_for_load_state("load", timeout=20000)
    except PlaywrightTimeoutError:
        _console.print("[yellow][!] 检索页加载超时，跳过该关键词。[/yellow]")
        return results
    except PlaywrightError as e:
        _console.print(f"[yellow][!] 检索页加载失败: {e}，跳过该关键词。[/yellow]")
        return results
    _check_captcha(page)

    # 输入关键词并发起检索，不捕获 KeyboardInterrupt
    with _console.status(
        "[bold magenta]少女祈祷中...[/bold magenta]",
        spinner="bouncingBar",
    ):
        page.fill("input.search-input", keyword)
        time.sleep(random.uniform(0.5, 1.5))
        page.click("input.search-btn")
        time.sleep(random.uniform(1, 2))
    _check_captcha(page)

    # v0.1.5: Progress 改为 bouncingBar + magenta 配色
    with Progress(
        SpinnerColumn(spinner_name="bouncingBar", style="bold magenta"),
        TextColumn("[bold magenta]{task.description}[/bold magenta]"),
        BarColumn(bar_width=36, style="magenta", complete_style="bright_magenta"),
        MofNCompleteColumn(),
        TextColumn("•"),
        TimeElapsedColumn(),
        console=_console,
        transient=False,
    ) as progress:
        task = progress.add_task(
            description=f"第 1 / {max_pages} 页",
            total=max_pages,
        )

        for current_page in range(1, max_pages + 1):
            # 整个循环体统一捕获，任意位置异常都能安全退出并保留已有数据
            # v0.1.5: 移除对 KeyboardInterrupt 的捕获，使其向上传递至 scrape_cnki
            try:
                progress.update(task, description=f"第 [bold]{current_page}[/bold] / {max_pages} 页")
                try:
                    page.wait_for_selector(
                        "table.result-table-list tbody tr", timeout=15000
                    )
                except PlaywrightTimeoutError:
                    had_captcha = _check_captcha(page, progress)
                    if had_captcha:
                        page.wait_for_selector(
                            "table.result-table-list tbody tr", timeout=15000
                        )
                    else:
                        progress.console.print(
                            f"[red][x] 第 {current_page} 页等待超时且未检测到验证码，跳过本页。[/red]"
                        )
                        progress.advance(task)
                        continue

                time.sleep(random.uniform(2, 5))
                _check_captcha(page, progress)

                rows = page.query_selector_all("table.result-table-list tbody tr")
                for row in rows:
                    title_el = row.query_selector("td.name a")
                    if title_el:
                        title = title_el.inner_text().strip()
                        results.append([title])
                        progress.console.print(f"  [green]→[/green] {title}")

                progress.advance(task)

                if current_page < max_pages:
                    next_btn = page.query_selector("a#PageNext")
                    if next_btn:
                        next_btn.click()
                        time.sleep(random.uniform(4, 8))
                        _check_captcha(page, progress)
                    else:
                        progress.console.print(
                            "[yellow][!] 没找到下一页按钮，可能已到最后一页。[/yellow]"
                        )
                        break

            except PlaywrightError:
                # 不调用 progress.stop()，让 with 块的 __exit__ 统一处理
                # 避免手动 stop + __exit__ 二次 stop 引发内部异常
                progress.console.print(
                    "\n[yellow][!] 检测到浏览器被手动关闭或强制中断，"
                    "正在为您安全中止并保存已抓取的数据...[/yellow]"
                )
                break

            # v0.1.6: 在循环内截获 Ctrl+C，results 中已有的数据完好保留
            # 同样不调用 progress.stop()——让 __exit__ 统一处理，消息移到 with 块外
            except KeyboardInterrupt:
                _stop_requested = True
                break

    # with Progress() 已正常退出，此处打印不会被进度条覆盖
    if _stop_requested:
        _console.print("[yellow][!] 用户中断，正在保存已抓取的数据...[/yellow]")

    return results


def scrape_cnki(keywords: list[str], max_pages: int, save_mode: str):
    """
    save_mode:
      'single'       -> 单关键词，保存为 cnki_titles_关键词.xlsx
      'multi_split'  -> 多关键词分文件保存
      'multi_merge'  -> 多关键词单文件多 Sheet 保存
    """
    all_results: dict[str, list] = {}

    # v0.1.6: 每次调用前重置标志位，防止同进程多轮运行时状态残留
    global _stop_requested
    _stop_requested = False

    with sync_playwright() as p:
        browser = None

        try:
            # v0.1.5: 浏览器启动包入 status 动画
            with _console.status(
                "[bold magenta]少女祈祷中...[/bold magenta]",
                spinner="bouncingBar",
            ):
                browser = p.chromium.launch(channel="msedge", headless=False)
            _console.print("[dim][*] 已启动 Microsoft Edge[/dim]")
        except Exception as _e1:
            _console.print(f"[yellow][!] Edge 启动失败 ({_e1})，尝试备用 Chromium...[/yellow]")
            try:
                with _console.status(
                    "[bold magenta]少女祈祷中...[/bold magenta]",
                    spinner="bouncingBar",
                ):
                    browser = p.chromium.launch(headless=False)
                _console.print("[dim][*] 已启动备用 Chromium 浏览器[/dim]")
            except Exception as _e2:
                if sys.platform == "win32":
                    _popup_error([
                        "==============================================",
                        " [错误] 浏览器启动失败！",
                        "----------------------------------------------",
                        " 程序找到了 Edge，但无法正常启动它。",
                        "",
                        " 可能原因：",
                        "   · Edge 浏览器文件损坏",
                        "   · 系统权限不足",
                        "   · 安全软件阻止了浏览器启动",
                        "",
                        " 建议：",
                        "   1. 重新安装 Microsoft Edge",
                        "      https://www.microsoft.com/zh-cn/edge/download",
                        "   2. 以管理员身份运行本程序",
                        "   3. 暂时关闭杀毒软件后重试",
                        "==============================================",
                    ])
                else:
                    _console.print(f"[red][FATAL] 浏览器启动失败: {_e2}[/red]")
                raise RuntimeError(f"浏览器启动彻底失败: {_e2}")

        try:
            context = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                )
            )
            page = context.new_page()

            # 预热搜索：消耗首次 302 重定向
            # v0.1.5: 将 _check_captcha 调用移到 status 块外，避免 input() 在 Live 上下文中运行
            dummy_keyword = "焊接"
            try:
                with _console.status(
                    "[bold magenta]少女祈祷中...[/bold magenta]",
                    spinner="bouncingBar",
                ):
                    page.goto("https://www.cnki.net/", timeout=30000)
                    page.wait_for_load_state("domcontentloaded", timeout=20000)
                _check_captcha(page)
                with _console.status(
                    "[bold magenta]少女祈祷中...[/bold magenta]",
                    spinner="bouncingBar",
                ):
                    page.goto("https://kns.cnki.net/kns8s/", timeout=30000)
                    page.wait_for_load_state("load", timeout=20000)
                    page.fill("input.search-input", dummy_keyword)
                    time.sleep(random.uniform(0.5, 1.5))
                    page.click("input.search-btn")
                    page.wait_for_selector("table.result-table-list tbody tr", timeout=15000)
                _check_captcha(page)
                _console.print("[dim][*] 预热完成，开始正式抓取。[/dim]")
            except (PlaywrightTimeoutError, PlaywrightError) as warmup_err:
                _console.print(f"[yellow][!] 预热搜索未完全成功 ({warmup_err})，继续正式抓取。[/yellow]")
            time.sleep(random.uniform(2, 4))

            for idx, keyword in enumerate(keywords):
                if idx > 0:
                    wait_sec = random.uniform(5, 8)
                    # v0.1.5: 冷却等待包入 status 动画
                    with _console.status(
                        f"[dim]少女祈祷中... 等待 {wait_sec:.1f} 秒[/dim]",
                        spinner="dots",
                    ):
                        time.sleep(wait_sec)

                # v0.1.6: results 在 try 外初始化，无论何种异常都能安全写入 all_results
                results = []
                try:
                    results = _scrape_keyword(page, keyword, max_pages)
                except PlaywrightTimeoutError as e:
                    _console.print(f"[red][x] 关键词「{keyword}」页面等待超时，跳过: {e}[/red]")
                except PlaywrightError as e:
                    _console.print(f"[yellow][!] 浏览器连接已断开，停止后续关键词抓取: {e}[/yellow]")
                    _stop_requested = True
                except KeyboardInterrupt:
                    # 导航阶段的 Ctrl+C（Progress 循环内的已在底层截获）
                    # results 此时为空，_stop_requested 将在此处置 True
                    _stop_requested = True

                # 无论正常返回、超时、浏览器断开、还是导航阶段中断，都记录当前结果
                all_results[keyword] = results

                if _stop_requested:
                    break

        # v0.1.5: 捕获 Ctrl+C，打印提示后不重新抛出
        # finally 负责关闭浏览器，执行流随后自然落入函数末尾的保存逻辑
        except KeyboardInterrupt:
            _console.print(
                "\n[bold yellow][!] 用户中断，正在保存已抓取的数据...[/bold yellow]"
            )
        except RuntimeError as e:
            _console.print(f"[red][x] 运行时错误: {e}[/red]")
        finally:
            if browser:
                # v0.1.6: 用 BaseException 兜底——Ctrl+C 是 BaseException 子类，
                # except Exception 无法捕获，会从 finally 逃出并跳过保存逻辑
                try:
                    browser.close()
                except BaseException:
                    pass

    if save_mode == "single":
        _save_single(keywords[0], all_results.get(keywords[0], []))
    elif save_mode == "multi_split":
        _save_multi_split(all_results)
    elif save_mode == "multi_merge":
        _save_multi_merge(all_results)


def _get_output_path(filename: str) -> str:
    try:
        real_desktop = get_real_desktop_path()
        os.makedirs(real_desktop, exist_ok=True)
        return os.path.join(real_desktop, filename)
    except OSError:
        return os.path.join(os.getcwd(), filename)


def _try_save_workbook(wb, filepath: str) -> bool:
    try:
        # v0.1.5: 写盘操作包入 status 动画，提供视觉反馈
        with _console.status(
            "[bold magenta]少女祈祷中...[/bold magenta]",
            spinner="bouncingBar",
        ):
            wb.save(filepath)
        return True
    except PermissionError:
        _console.print(f"\n[red][x] 文件保存失败：没有写入权限！[/red]")
        _console.print(f"    目标路径：{filepath}")
        _console.print(f"    请确认桌面文件夹未被锁定，或关闭已打开的同名 Excel 文件。")
        return False
    except OSError as save_err:
        _console.print(f"\n[red][x] 文件保存失败：{save_err}[/red]")
        fallback = os.path.join(os.getcwd(), os.path.basename(filepath))
        _console.print(f"    尝试保存到程序目录：{fallback}")
        try:
            wb.save(fallback)
            _console.print(f"    已保存至备用路径：{fallback}")
            return True
        except OSError as fb_err:
            _console.print(f"[red][x] 备用路径也保存失败：{fb_err}[/red]")
            return False


def _save_single(keyword: str, results: list):
    if not results:
        _console.print("[yellow][!] 未抓取到任何数据，不生成文件。[/yellow]")
        return

    clean_keyword = _sanitize_name(keyword)
    filepath = _get_output_path(f"cnki_titles_{clean_keyword}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "论文标题"
    ws.append(["论文标题"])
    for row in results:
        ws.append(row)

    if _try_save_workbook(wb, filepath):
        _console.print("\n" + "═" * 50)
        _console.print(f"[bold green][*] 共抓取 {len(results)} 条数据。[/bold green]")
        _console.print(f"[*] 文件已保存至：")
        _console.print(f"    [bold]>>> {os.path.abspath(filepath)} <<<[/bold]")
        _console.print("═" * 50 + "\n")


def _save_multi_split(all_results: dict[str, list]):
    total = 0
    saved_files = []
    for keyword, results in all_results.items():
        if not results:
            _console.print(f"[yellow][!] 关键词「{keyword}」无数据，跳过生成文件。[/yellow]")
            continue
        clean_keyword = _sanitize_name(keyword)
        filepath = _get_output_path(f"cnki_titles_{clean_keyword}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "论文标题"
        ws.append(["论文标题"])
        for row in results:
            ws.append(row)
        if _try_save_workbook(wb, filepath):
            saved_files.append((keyword, len(results), os.path.abspath(filepath)))
            total += len(results)

    _console.print("\n" + "═" * 50)
    _console.print(
        f"[bold green][*] 全部抓取完毕，共 {total} 条数据，生成 {len(saved_files)} 个文件：[/bold green]"
    )
    for kw, cnt, path in saved_files:
        _console.print(f"  · [cyan][{kw}][/cyan] {cnt} 条  ->  {path}")
    _console.print("═" * 50 + "\n")


def _save_multi_merge(all_results: dict[str, list]):
    if not any(len(v) > 0 for v in all_results.values()):
        _console.print("[yellow][!] 所有关键词均未抓取到数据，不生成文件。[/yellow]")
        return

    filepath = _get_output_path("cnki_titles_多词汇总.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total = 0
    # 记录已使用的 sheet 名，截断后若重复则追加 _1/_2 ... 保证唯一
    used_sheet_names: set[str] = set()
    for keyword, results in all_results.items():
        clean_keyword = _sanitize_name(keyword)
        base_name = clean_keyword[:31]
        sheet_name = base_name
        counter = 1
        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            sheet_name = base_name[:31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append(["论文标题"])
        for row in results:
            ws.append(row)
        total += len(results)

    if _try_save_workbook(wb, filepath):
        _console.print("\n" + "═" * 50)
        _console.print(f"[bold green][*] 全部抓取完毕，共 {total} 条数据。[/bold green]")
        _console.print(f"[*] 已合并保存至：")
        _console.print(f"    [bold]>>> {os.path.abspath(filepath)} <<<[/bold]")
        for kw, results in all_results.items():
            _console.print(f"  · Sheet [cyan][{kw}][/cyan]：{len(results)} 条")
        _console.print("═" * 50 + "\n")


if __name__ == "__main__":
    try:
        if sys.platform == "win32":
            os.system("cls")

        # v0.1.5: 启动横幅
        _console.print("=" * 50)
        _console.print("  CNKI_Bug_dev  |  copyright by Kaffu_Alcaid")
        _console.print("  Version 0.1.5")
        _console.print("=" * 50)
        _console.print("  本软件用于抓取中国知网的论文标题\n")
        _console.print("按 Ctrl+C 可随时中断并保存已抓取数据")
        _console.print("それは，幾千の夜を舞う、さくらと少女たちの物語ーーー")
        _console.print("祈祷着今后的你的人生，永远都有幸福的“魔法”相伴")

        check_env()

        while True:
            try:
                print("\n请选择抓取模式：")
                print("  1 -> 单关键词模式")
                print("  2 -> 多关键词模式")
                mode_input = input("请输入选项（1 或 2）: ").strip()
                if mode_input not in ("1", "2"):
                    print("[!] 无效选项，程序退出。")
                    sys.exit(0)

                keywords = []
                if mode_input == "1":
                    word = input("\n请输入你要搜索的关键词: ").strip()
                    if not word:
                        print("[!] 关键词不能为空，程序退出。")
                        sys.exit(0)
                    keywords = [word]
                    save_mode = "single"
                else:
                    print("\n请依次输入关键词，每输入一个按回车确认；直接按回车结束输入：")
                    while True:
                        word = input("  关键词: ").strip()
                        if not word:
                            break
                        keywords.append(word)
                    if not keywords:
                        print("[!] 未输入任何关键词，程序退出。")
                        sys.exit(0)
                    print(f"\n[*] 共确认 {len(keywords)} 个关键词：{keywords}")

                    print("\n请选择保存方式：")
                    print("  1 -> 分文件保存（每个关键词独立生成一个 Excel）")
                    print("  2 -> 单文件多 Sheet 保存（所有关键词汇总到一个 Excel）")
                    save_input = input("请输入选项（1 或 2）: ").strip()
                    if save_input == "1":
                        save_mode = "multi_split"
                    elif save_input == "2":
                        save_mode = "multi_merge"
                    else:
                        print("[!] 无效选项，程序退出。")
                        sys.exit(0)

                # 页数输入独立内层循环，ValueError 在此消化，不影响已选模式和关键词
                while True:
                    try:
                        user_input_pages = input("\n请输入想抓取的总页数（纯数字，值不要太大）: ").strip()
                        target_pages = int(user_input_pages)
                        if target_pages <= 0:
                            print("  [!] 页数必须大于 0，请重新输入。")
                            continue
                    except ValueError:
                        print("  [!] 错误：页数请输入【纯数字】，例如 3 或 10，请重新输入。")
                        continue

                    if target_pages > 20:
                        _console.print(
                            f"\n[yellow][!] 您输入的页数较大（{target_pages}页），"
                            f"预计将耗时较长，且容易触发知网反爬验证。[/yellow]"
                        )
                        confirm = input("确定要继续吗？(y/n): ").strip().lower()
                        if confirm == "y":
                            break
                        else:
                            continue
                    else:
                        break

                scrape_cnki(keywords, max_pages=target_pages, save_mode=save_mode)

                again = input("\n[*] 本轮抓取已完成！是否清屏并开始新一轮抓取？(y/n): ").strip().lower()
                if again == "y":
                    if sys.platform == "win32":
                        os.system("cls")
                    continue
                else:
                    # v0.1.5: 退出提示
                    _console.print("\n[bold green]感谢使用 CNKIBug，再见！[/bold green]")
                    break

            except RuntimeError as e:
                print(f"\n[!] {e}")
                retry = input("是否返回主菜单重试？(y/n): ").strip().lower()
                if retry == "y":
                    continue
                else:
                    break

            except Exception as ex:
                print("\n" + "!" * 40)
                print(f"  程序遇到未知错误: {ex}")
                print("!" * 40)
                break

    except KeyboardInterrupt:
        print("\n[*] 用户中断，程序退出。")
    finally:
        input("\n按 [回车键 Enter] 退出程序...")
