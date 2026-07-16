import csv
import os

import openpyxl

from cnkibug.fileio import exporter
from cnkibug.fileio.exporter import (
    _build_single_sheet_workbook,
    _sanitize_name,
    _try_save_workbook,
    save_all,
)


# ============ 原有用例（回归） ============
def test_sanitize_name_replaces_filename_and_sheet_illegal_chars():
    assert _sanitize_name('a/b:c*?"<>|[x]') == "a_b_c_______x_"


def test_build_single_sheet_workbook_headers_and_rows():
    detail_url = "https://kns.cnki.net/detail/1"
    wb = _build_single_sheet_workbook([["标题", "作者", "来源", "2026-01-01", detail_url]])
    ws = wb.active

    assert ws is not None
    assert ws.title == "论文标题"
    assert [cell.value for cell in ws[1]] == ["论文标题", "作者", "来源", "发表日期", "详情链接"]
    assert [cell.value for cell in ws[2]] == ["标题", "作者", "来源", "2026-01-01", detail_url]
    assert ws["E2"].hyperlink.target == detail_url


def test_build_workbook_inserts_citation_before_detail_url():
    detail_url = "https://kns.cnki.net/detail/1"
    citation = "[1] 示例引文[J]. 测试期刊,2026."
    wb = _build_single_sheet_workbook(
        [["标题", "作者", "来源", "2026-01-01", detail_url, citation]],
        include_citation=True,
    )
    ws = wb.active

    assert ws is not None
    assert [cell.value for cell in ws[1]] == [
        "论文标题",
        "作者",
        "来源",
        "发表日期",
        "引用格式",
        "详情链接",
    ]
    assert [cell.value for cell in ws[2]] == [
        "标题",
        "作者",
        "来源",
        "2026-01-01",
        citation,
        detail_url,
    ]
    assert ws["F2"].hyperlink.target == detail_url


# ============ _sanitize_name 新边界（A11） ============
def test_sanitize_name_pure_dots_and_empty_fallback():
    # strip 后为空（纯点 / 纯空白 / 空串）→ 兜底默认名
    assert _sanitize_name("..") == "untitled"
    assert _sanitize_name("...") == "untitled"
    assert _sanitize_name("   ") == "untitled"
    assert _sanitize_name("") == "untitled"


def test_sanitize_name_all_illegal_chars_become_legal_underscores():
    # 全非法字符被替换为下划线串，本身是合法文件名，不走兜底
    assert _sanitize_name("///") == "___"


def test_sanitize_name_truncates_long_input():
    assert _sanitize_name("x" * 80) == "x" * 50


def test_sanitize_name_keeps_normal_keyword():
    assert _sanitize_name("焊接 316L") == "焊接 316L"


# ============ helpers ============
def _patch_desktop(monkeypatch, tmp_path):
    """把导出目录重定向到临时目录，避免污染真实桌面。"""
    monkeypatch.setattr(exporter, "get_real_desktop_path", lambda: str(tmp_path))


def _load(path):
    return openpyxl.load_workbook(path)


# ============ save_all: single ============
def test_save_all_single_writes_file(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    data = [["t1", "a1", "s1", "2026-01-01"], ["t2", "a2", "s2", "2026-02-02"]]
    result = save_all("single", ["焊接"], {"焊接": data}, "TS")

    files = list(tmp_path.glob("cnki_titles_焊接_TS.xlsx"))
    assert len(files) == 1
    assert result.attempted == 1
    assert result.failed == 0
    assert result.saved_paths == [str(files[0].resolve())]
    ws = _load(files[0]).active
    assert [c.value for c in ws[1]] == ["论文标题", "作者", "来源", "发表日期", "详情链接"]
    assert ws.max_row == 3  # 表头 + 2 行数据


def test_save_all_single_no_data_skips_file(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    result = save_all("single", ["焊接"], {"焊接": []}, "TS")
    assert result.attempted == 0
    assert result.failed == 0
    assert result.saved_paths == []
    assert list(tmp_path.glob("*.xlsx")) == []


def test_save_all_single_csv_writes_keyword_column(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    data = [["标题", "作者", "来源", "2026-01-01", "https://example.test/1"]]

    result = save_all("single_csv", ["焊接"], {"焊接": data}, "TS")

    path = tmp_path / "cnki_titles_焊接_TS.csv"
    assert result.saved_paths == [str(path.resolve())]
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    assert rows == [
        ["keyword", "title", "authors", "source", "publication_date", "detail_url"],
        ["焊接", "标题", "作者", "来源", "2026-01-01", "https://example.test/1"],
    ]


def test_save_all_single_csv_inserts_citation_before_detail_url(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    data = [[
        "标题",
        "作者",
        "来源",
        "2026-01-01",
        "https://example.test/1",
        "[1] 示例引文",
    ]]

    save_all(
        "single_csv",
        ["焊接"],
        {"焊接": data},
        "TS",
        include_citation=True,
    )

    path = tmp_path / "cnki_titles_焊接_TS.csv"
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    assert rows == [
        [
            "keyword",
            "title",
            "authors",
            "source",
            "publication_date",
            "citation",
            "detail_url",
        ],
        [
            "焊接",
            "标题",
            "作者",
            "来源",
            "2026-01-01",
            "[1] 示例引文",
            "https://example.test/1",
        ],
    ]


# ============ multi_split：每词一文件 ============
def test_save_all_multi_split_one_file_per_keyword(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    all_results = {
        "焊接": [["t", "a", "s", "d"]],
        "增材": [["t2", "a2", "s2", "d2"], ["t3", "a3", "s3", "d3"]],
    }
    save_all("multi_split", list(all_results), all_results, "TS")

    assert (tmp_path / "cnki_titles_焊接_TS.xlsx").exists()
    f2 = tmp_path / "cnki_titles_增材_TS.xlsx"
    assert f2.exists()
    assert _load(f2).active.max_row == 3


def test_save_all_multi_split_avoids_sanitized_name_collision(monkeypatch, tmp_path, caplog):
    _patch_desktop(monkeypatch, tmp_path)
    all_results = {
        "AI/ML": [["first", "a", "s", "d"]],
        "AI:ML": [["second", "a", "s", "d"]],
    }

    result = save_all("multi_split", list(all_results), all_results, "TS")

    first = tmp_path / "cnki_titles_AI_ML_TS.xlsx"
    second = tmp_path / "cnki_titles_AI_ML_2_TS.xlsx"
    assert first.exists()
    assert second.exists()
    assert _load(first).active["A2"].value == "first"
    assert _load(second).active["A2"].value == "second"
    assert len(result.saved_paths) == 2
    assert "分文件保存名冲突" in caplog.text


def test_save_all_multi_split_skips_empty_keyword(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    all_results = {"有": [["t", "a", "s", "d"]], "无": []}
    save_all("multi_split", ["有", "无"], all_results, "TS")

    assert (tmp_path / "cnki_titles_有_TS.xlsx").exists()
    assert not (tmp_path / "cnki_titles_无_TS.xlsx").exists()


# ============ multi_merge：单文件多 Sheet ============
def test_save_all_multi_merge_one_file_multi_sheet(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    all_results = {
        "焊接": [["t", "a", "s", "d"]],
        "增材": [["t2", "a2", "s2", "d2"]],
    }
    save_all("multi_merge", list(all_results), all_results, "TS")

    files = list(tmp_path.glob("cnki_titles_多词汇总_TS.xlsx"))
    assert len(files) == 1
    wb = _load(files[0])
    assert wb.sheetnames == ["焊接", "增材"]
    assert [c.value for c in wb["焊接"][1]] == ["论文标题", "作者", "来源", "发表日期", "详情链接"]


def test_multi_merge_sheet_name_truncated_and_deduped(monkeypatch, tmp_path):
    """Sheet 名超 31 字截断 + 截断后撞名加后缀去重（Excel 硬上限）。"""
    _patch_desktop(monkeypatch, tmp_path)
    k1 = "X" * 35
    k2 = "X" * 31 + "YYYY"  # 前 31 字符与 k1 相同 → 截断后撞名
    all_results = {k1: [["t", "a", "s", "d"]], k2: [["t2", "a2", "s2", "d2"]]}
    save_all("multi_merge", [k1, k2], all_results, "TS")

    wb = _load(tmp_path / "cnki_titles_多词汇总_TS.xlsx")
    names = wb.sheetnames
    assert len(names) == 2
    assert all(len(n) <= 31 for n in names)  # 不超 Excel 31 字上限
    assert len(set(names)) == 2              # 去重成功，未互相覆盖
    assert names[0] == "X" * 31


def test_save_all_multi_merge_all_empty_no_file(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    save_all("multi_merge", ["a", "b"], {"a": [], "b": []}, "TS")
    assert list(tmp_path.glob("*.xlsx")) == []


def test_save_all_multi_csv_writes_flat_utf8_file(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    all_results = {
        "焊接": [["标题一", "作者甲", "来源甲", "2026-01-01", "https://example.test/1"]],
        "增材": [["标题,二", "作者乙", "来源乙", "", ""]],
    }

    result = save_all("multi_csv", list(all_results), all_results, "TS")

    path = tmp_path / "cnki_titles_多词汇总_TS.csv"
    assert result.saved_paths == [str(path.resolve())]
    with path.open(encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    assert rows == [
        ["keyword", "title", "authors", "source", "publication_date", "detail_url"],
        ["焊接", "标题一", "作者甲", "来源甲", "2026-01-01", "https://example.test/1"],
        ["增材", "标题,二", "作者乙", "来源乙", "", ""],
    ]


def test_save_all_multi_csv_skips_empty_results(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)

    result = save_all("multi_csv", ["空"], {"空": []}, "TS")

    assert result.attempted == 0
    assert list(tmp_path.glob("*.csv")) == []


# ============ PermissionError 回退到程序目录（核心防丢逻辑） ============
def test_try_save_workbook_falls_back_to_cwd_on_permission_error(monkeypatch, tmp_path):
    monkeypatch.chdir(tmp_path)  # 让回退落点 = 临时目录，不污染真实 cwd
    wb = _build_single_sheet_workbook([["t", "a", "s", "d"]])
    target = os.path.join(str(tmp_path), "locked", "out.xlsx")

    real_save = wb.save
    calls = []

    def fake_save(path):
        calls.append(str(path))
        if len(calls) == 1:
            raise PermissionError("file is open in Excel")  # 模拟桌面文件被占用
        return real_save(path)

    monkeypatch.setattr(wb, "save", fake_save)
    saved = _try_save_workbook(wb, target)

    assert saved is not None
    assert os.path.basename(saved) == "out.xlsx"
    assert os.path.dirname(os.path.abspath(saved)) == str(tmp_path)
    assert os.path.exists(saved)


def test_save_all_reports_failed_save(monkeypatch, tmp_path):
    _patch_desktop(monkeypatch, tmp_path)
    monkeypatch.setattr(exporter, "_try_save_workbook", lambda wb, filepath, **kwargs: None)

    result = save_all(
        "single",
        ["焊接"],
        {"焊接": [["t", "a", "s", "d"]]},
        "TS",
    )

    assert result.attempted == 1
    assert result.failed == 1
    assert result.saved_paths == []
    assert list(tmp_path.glob("*.xlsx")) == []
