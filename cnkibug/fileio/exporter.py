import csv
import logging
import os
import re
from dataclasses import dataclass, field

import openpyxl

from ..cnki.models import record_article_details, record_citation
from .paths import get_real_desktop_path

_CSV_HEADERS = ["keyword", "title", "authors", "source", "publication_date", "detail_url"]
_logger = logging.getLogger("cnkibug.exporter")


@dataclass
class SavedFile:
    path: str
    keyword: str = ""
    record_count: int = 0


@dataclass
class SaveResult:
    attempted: int = 0
    saved_paths: list[str] = field(default_factory=list)
    files: list[SavedFile] = field(default_factory=list)
    failed: int = 0
    keyword_txt_path: str | None = None
    keyword_txt_failed: bool = False

    def record(
        self,
        saved_path: str | None,
        *,
        keyword: str = "",
        record_count: int = 0,
    ) -> None:
        self.attempted += 1
        if saved_path:
            self.saved_paths.append(saved_path)
            self.files.append(SavedFile(saved_path, keyword, record_count))
        else:
            self.failed += 1


def _sanitize_name(text: str) -> str:
    """清洗为合法文件名/Sheet 名。

    除替换非法字符外，再做：去首尾空白与点（避免纯点名 . / ..）、限长到 50 字
    （防止文件名突破 Windows 260 路径上限导致写盘失败）、空结果给默认名。
    """
    cleaned = re.sub(r'[\\/:*?"<>|\[\]]', '_', text)
    cleaned = cleaned.strip().strip('.').strip()[:50].rstrip('. ')
    return cleaned or "untitled"


def _get_output_path(
    filename: str,
    output_dir: str | os.PathLike | None = None,
) -> str:
    try:
        target_dir = os.fspath(output_dir) if output_dir is not None else get_real_desktop_path()
        os.makedirs(target_dir, exist_ok=True)
        return os.path.join(target_dir, filename)
    except OSError:
        return os.path.join(os.getcwd(), filename)


def _try_save_fallback(
    wb,
    filepath: str,
    save_err: OSError,
    log_save_path: bool,
) -> str | None:
    fallback = os.path.join(os.getcwd(), os.path.basename(filepath))
    if log_save_path:
        _logger.warning("文件保存失败，尝试备用路径: target=%s fallback=%s error=%s", filepath, fallback, save_err)
    else:
        _logger.warning("文件保存失败，尝试备用路径: error=%s", save_err)

    try:
        wb.save(fallback)
        saved_path = os.path.abspath(fallback)
        _log_save_success(saved_path, log_save_path, "fallback")
        return saved_path
    except OSError as fb_err:
        _logger.error("备用路径保存失败: %s", fb_err)
        return None


def _log_save_success(saved_path: str, log_save_path: bool, save_type: str) -> None:
    if log_save_path:
        _logger.info("文件保存成功: type=%s path=%s", save_type, saved_path)
    else:
        _logger.info("文件保存成功: type=%s", save_type)


def _try_save_workbook(
    wb,
    filepath: str,
    *,
    log_save_path: bool = True,
    save_type: str = "final",
) -> str | None:
    try:
        wb.save(filepath)
        saved_path = os.path.abspath(filepath)
        _log_save_success(saved_path, log_save_path, save_type)
        return saved_path
    except OSError as save_err:
        return _try_save_fallback(wb, filepath, save_err, log_save_path)


def _export_headers(
    include_citation: bool,
    include_details: bool = False,
) -> list[str]:
    headers = ["论文标题", "作者", "来源", "发表日期"]
    if include_details:
        headers.extend(("论文关键词", "摘要"))
    if include_citation:
        headers.append("引用格式")
    headers.append("详情链接")
    return headers


def _export_record(
    record: list,
    include_citation: bool,
    include_details: bool = False,
) -> list:
    values = list(record[:5])
    values.extend([""] * (5 - len(values)))
    exported = list(values[:4])
    if include_details:
        keywords, abstract = record_article_details(record, include_citation)
        exported.extend((
            "；".join(item.strip() for item in keywords.splitlines() if item.strip()),
            _clean_cell_text(abstract),
        ))
    if include_citation:
        exported.append(record_citation(record, include_citation))
    exported.append(values[4])
    return exported


def _clean_cell_text(value: str) -> str:
    cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))
    if len(cleaned) > 32767:
        _logger.warning("导出字段超过 Excel 单元格上限，已截断: length=%d", len(cleaned))
        return cleaned[:32767]
    return cleaned


def _build_single_sheet_workbook(
    results: list,
    include_citation: bool = False,
    include_details: bool = False,
):
    """构建单 Sheet 工作簿（single / multi_split 共用）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "论文标题"
    ws.append(_export_headers(include_citation, include_details))
    _append_records(ws, results, include_citation, include_details)
    return wb


def _append_records(
    ws,
    results: list,
    include_citation: bool = False,
    include_details: bool = False,
) -> None:
    for row in results:
        ws.append(_export_record(row, include_citation, include_details))
        if len(row) > 4 and str(row[4]).strip():
            link_column = len(_export_headers(include_citation, include_details))
            link_cell = ws.cell(row=ws.max_row, column=link_column)
            link_cell.hyperlink = str(row[4]).strip()
            link_cell.style = "Hyperlink"


def _save_single(
    keyword: str,
    results: list,
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
):
    save_result = SaveResult()
    if not results:
        return save_result

    clean_keyword = _sanitize_name(keyword)
    filepath = _get_output_path(f"cnki_titles_{clean_keyword}_{ts}.xlsx", output_dir)
    wb = _build_single_sheet_workbook(results, include_citation, include_details)

    saved_path = _try_save_workbook(
        wb,
        filepath,
        log_save_path=log_save_path,
        save_type=save_type,
    )
    save_result.record(saved_path, keyword=keyword, record_count=len(results))
    return save_result


def _save_multi_split(
    all_results: dict[str, list],
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
):
    save_result = SaveResult()
    used_names: set[str] = set()
    for keyword, results in all_results.items():
        if not results:
            continue

        base_keyword = _sanitize_name(keyword)
        clean_keyword = base_keyword
        counter = 2
        while clean_keyword.casefold() in used_names:
            suffix = f"_{counter}"
            clean_keyword = f"{base_keyword[:50 - len(suffix)]}{suffix}"
            counter += 1
        if clean_keyword != base_keyword:
            _logger.warning(
                "分文件保存名冲突，已自动添加后缀: collision_index=%d",
                counter - 1,
            )
        used_names.add(clean_keyword.casefold())
        filepath = _get_output_path(
            f"cnki_titles_{clean_keyword}_{ts}.xlsx",
            output_dir,
        )
        wb = _build_single_sheet_workbook(results, include_citation, include_details)
        saved_path = _try_save_workbook(
            wb,
            filepath,
            log_save_path=log_save_path,
            save_type=save_type,
        )
        save_result.record(saved_path, keyword=keyword, record_count=len(results))
    return save_result


def _save_multi_merge(
    all_results: dict[str, list],
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
):
    save_result = SaveResult()
    if not any(len(v) > 0 for v in all_results.values()):
        return save_result

    filepath = _get_output_path(f"cnki_titles_多词汇总_{ts}.xlsx", output_dir)
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # noqa

    total = 0

    used_sheet_names: set[str] = set()
    for keyword, results in all_results.items():
        if not results:
            continue
        clean_keyword = _sanitize_name(keyword)
        base_name = clean_keyword[:31]
        sheet_name = base_name
        counter = 1
        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            sheet_name = base_name[:31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append(_export_headers(include_citation, include_details))
        _append_records(ws, results, include_citation, include_details)
        total += len(results)

    saved_path = _try_save_workbook(
        wb,
        filepath,
        log_save_path=log_save_path,
        save_type=save_type,
    )
    save_result.record(saved_path, record_count=total)
    return save_result


def _write_multi_csv(
    filepath: str,
    all_results: dict[str, list],
    include_citation: bool = False,
    include_details: bool = False,
) -> None:
    with open(filepath, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        headers = list(_CSV_HEADERS)
        if include_details:
            headers[5:5] = ["paper_keywords", "abstract"]
        if include_citation:
            headers.insert(7 if include_details else 5, "citation")
        writer.writerow(headers)
        for keyword, records in all_results.items():
            for record in records:
                values = _export_record(record, include_citation, include_details)
                writer.writerow([keyword, *values])


def _try_save_csv(
    filepath: str,
    all_results: dict[str, list],
    include_citation: bool = False,
    include_details: bool = False,
    *,
    log_save_path: bool = True,
    save_type: str = "final",
) -> str | None:
    try:
        _write_multi_csv(filepath, all_results, include_citation, include_details)
        saved_path = os.path.abspath(filepath)
        _log_save_success(saved_path, log_save_path, save_type)
        return saved_path
    except OSError as save_err:
        fallback = os.path.join(os.getcwd(), os.path.basename(filepath))
        if log_save_path:
            _logger.warning(
                "CSV 保存失败，尝试备用路径: target=%s fallback=%s error=%s",
                filepath,
                fallback,
                save_err,
            )
        else:
            _logger.warning("CSV 保存失败，尝试备用路径: error=%s", save_err)
        try:
            _write_multi_csv(fallback, all_results, include_citation, include_details)
            saved_path = os.path.abspath(fallback)
            _log_save_success(saved_path, log_save_path, "fallback")
            return saved_path
        except OSError as fallback_err:
            _logger.error("CSV 备用路径保存失败: %s", fallback_err)
            return None


def _save_multi_csv(
    all_results: dict[str, list],
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
) -> SaveResult:
    save_result = SaveResult()
    total = sum(len(records) for records in all_results.values())
    if total == 0:
        return save_result

    filepath = _get_output_path(f"cnki_titles_多词汇总_{ts}.csv", output_dir)
    saved_path = _try_save_csv(
        filepath,
        all_results,
        include_citation,
        include_details,
        log_save_path=log_save_path,
        save_type=save_type,
    )
    save_result.record(saved_path, record_count=total)
    return save_result


def _save_single_csv(
    keyword: str,
    results: list,
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
) -> SaveResult:
    save_result = SaveResult()
    if not results:
        return save_result

    clean_keyword = _sanitize_name(keyword)
    filepath = _get_output_path(f"cnki_titles_{clean_keyword}_{ts}.csv", output_dir)
    saved_path = _try_save_csv(
        filepath,
        {keyword: results},
        include_citation,
        include_details,
        log_save_path=log_save_path,
        save_type=save_type,
    )
    save_result.record(saved_path, keyword=keyword, record_count=len(results))
    return save_result


def _save_keyword_txt(
    result: SaveResult,
    all_results: dict[str, list],
    ts: str,
    include_citation: bool,
    log_save_path: bool,
    save_type: str,
    output_dir: str | os.PathLike | None = None,
) -> None:
    lines = []
    for records in all_results.values():
        for record in records:
            keywords, _ = record_article_details(record, include_citation)
            lines.extend(item.strip() for item in keywords.splitlines() if item.strip())
    if not lines:
        return

    filepath = _get_output_path(f"cnki_paper_keywords_{ts}.txt", output_dir)
    try:
        _write_keyword_txt(filepath, lines)
        result.keyword_txt_path = os.path.abspath(filepath)
        _log_save_success(result.keyword_txt_path, log_save_path, save_type)
        return
    except OSError as save_error:
        fallback = os.path.join(os.getcwd(), os.path.basename(filepath))
        if log_save_path:
            _logger.warning(
                "关键词 TXT 保存失败，尝试备用路径: target=%s fallback=%s error=%s",
                filepath,
                fallback,
                save_error,
            )
        else:
            _logger.warning("关键词 TXT 保存失败，尝试备用路径: error=%s", save_error)
    try:
        _write_keyword_txt(fallback, lines)
        result.keyword_txt_path = os.path.abspath(fallback)
        _log_save_success(result.keyword_txt_path, log_save_path, "fallback")
    except OSError as fallback_error:
        result.keyword_txt_failed = True
        _logger.error("关键词 TXT 备用路径保存失败: %s", fallback_error)


def _write_keyword_txt(filepath: str, lines: list[str]) -> None:
    with open(filepath, "w", encoding="utf-8-sig", newline="\n") as file:
        for line in lines:
            file.write(line + "\n")


def save_all(
    save_mode: str,
    keywords: list[str],
    all_results: dict[str, list],
    ts: str,
    include_citation: bool = False,
    include_details: bool = False,
    detail_txt_export: bool = False,
    *,
    output_dir: str | os.PathLike | None = None,
    log_save_path: bool = True,
    save_type: str = "final",
) -> SaveResult:
    """Persist the current result snapshot without producing UI output."""
    result = SaveResult()
    if save_mode == "single":
        if keywords:
            result = _save_single(
                keywords[0],
                all_results.get(keywords[0], []),
                ts,
                include_citation,
                include_details,
                output_dir=output_dir,
                log_save_path=log_save_path,
                save_type=save_type,
            )
    elif save_mode == "single_csv":
        if keywords:
            result = _save_single_csv(
                keywords[0],
                all_results.get(keywords[0], []),
                ts,
                include_citation,
                include_details,
                output_dir=output_dir,
                log_save_path=log_save_path,
                save_type=save_type,
            )
    elif save_mode == "multi_split":
        result = _save_multi_split(
            all_results,
            ts,
            include_citation,
            include_details,
            output_dir=output_dir,
            log_save_path=log_save_path,
            save_type=save_type,
        )
    elif save_mode == "multi_merge":
        result = _save_multi_merge(
            all_results,
            ts,
            include_citation,
            include_details,
            output_dir=output_dir,
            log_save_path=log_save_path,
            save_type=save_type,
        )
    elif save_mode == "multi_csv":
        result = _save_multi_csv(
            all_results,
            ts,
            include_citation,
            include_details,
            output_dir=output_dir,
            log_save_path=log_save_path,
            save_type=save_type,
        )
    if include_details and detail_txt_export:
        _save_keyword_txt(
            result,
            all_results,
            ts,
            include_citation,
            log_save_path,
            save_type,
            output_dir,
        )
    return result
