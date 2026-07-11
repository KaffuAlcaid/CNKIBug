import os
import re
import logging
from dataclasses import dataclass, field

import openpyxl

from .ui import _console
from .environment import get_real_desktop_path
from .runtime import get_config

# 三种保存模式共用的表头，与 scraper 抓取的列顺序一一对应：
# [论文标题, 作者, 来源, 发表日期, 详情链接]
_HEADERS = ["论文标题", "作者", "来源", "发表日期", "详情链接"]
_logger = logging.getLogger("cnkibug.exporter")


@dataclass
class SaveResult:
    attempted: int = 0
    saved_paths: list[str] = field(default_factory=list)
    failed: int = 0

    def record(self, saved_path: str | None) -> None:
        self.attempted += 1
        if saved_path:
            self.saved_paths.append(saved_path)
        else:
            self.failed += 1


def _sanitize_name(text: str) -> str:
    """清洗为合法文件名/Sheet 名。

    除替换非法字符外，再做：去首尾空白与点（避免纯点名 . / ..）、限长到 50 字
    （防止文件名突破 Windows 260 路径上限导致写盘失败）、空结果给默认名。
    """
    cleaned = re.sub(r'[\\/:*?"<>|\[\]]', '_', text)
    cleaned = cleaned.strip().strip('.').strip()[:50].rstrip('. ')
    return cleaned or "untitled"


def _get_output_path(filename: str) -> str:
    try:
        real_desktop = get_real_desktop_path()
        os.makedirs(real_desktop, exist_ok=True)
        return os.path.join(real_desktop, filename)
    except OSError:
        return os.path.join(os.getcwd(), filename)


def _try_save_fallback(wb, filepath: str, save_err: OSError, announce: bool) -> str | None:
    fallback = os.path.join(os.getcwd(), os.path.basename(filepath))
    if _log_save_path_enabled():
        _logger.warning("文件保存失败，尝试备用路径: target=%s fallback=%s error=%s", filepath, fallback, save_err)
    else:
        _logger.warning("文件保存失败，尝试备用路径: error=%s", save_err)
    if announce:
        _console.print(f"\n[red][x] 文件保存失败：{save_err}[/red]")
        _console.print("    可能原因：桌面不可写、同名 Excel 正在打开，或目录权限不足。")
        _console.print(f"    正在尝试保存到备用位置：{fallback}")

    try:
        wb.save(fallback)
        saved_path = os.path.abspath(fallback)
        _log_save_success(saved_path, announce)
        if announce:
            _console.print(f"    已保存至备用位置：{saved_path}")
        return saved_path
    except OSError as fb_err:
        _logger.error("备用路径保存失败: %s", fb_err)
        if announce:
            _console.print(f"[red][x] 备用路径也保存失败：{fb_err}[/red]")
            _console.print("[yellow]请关闭已打开的同名 Excel 文件，并检查桌面或程序目录写入权限。[/yellow]")
        return None


def _log_save_path_enabled() -> bool:
    return bool(get_config().get("log_save_path", True))


def _log_save_success(saved_path: str, announce: bool) -> None:
    save_type = "final" if announce else "incremental"
    if _log_save_path_enabled():
        _logger.info("文件保存成功: type=%s path=%s", save_type, saved_path)
    else:
        _logger.info("文件保存成功: type=%s", save_type)


def _try_save_workbook(wb, filepath: str, announce: bool = True) -> str | None:
    """写盘并返回实际保存路径。

    announce=False 时静默（不显示 spinner、不打印失败提示），供每抓完一个
    关键词的增量落盘使用；失败由调用方记日志、最终保存时再提示。
    """
    try:
        if announce:
            with _console.status(
                "[bold magenta]少女祈祷中...[/bold magenta]",
                spinner="bouncingBar",
            ):
                wb.save(filepath)
        else:
            wb.save(filepath)
        saved_path = os.path.abspath(filepath)
        _log_save_success(saved_path, announce)
        return saved_path
    except OSError as save_err:
        return _try_save_fallback(wb, filepath, save_err, announce)


def _build_single_sheet_workbook(results: list):
    """构建单 Sheet 工作簿（single / multi_split 共用）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "论文标题"
    ws.append(_HEADERS)
    _append_records(ws, results)
    return wb


def _append_records(ws, results: list) -> None:
    for row in results:
        ws.append(row)
        if len(row) > 4 and str(row[4]).strip():
            link_cell = ws.cell(row=ws.max_row, column=5)
            link_cell.hyperlink = str(row[4]).strip()
            link_cell.style = "Hyperlink"


def _save_single(keyword: str, results: list, ts: str, announce: bool):
    save_result = SaveResult()
    if not results:
        if announce:
            _console.print("[yellow][!] 未抓取到任何数据，不生成文件。[/yellow]")
        return save_result

    clean_keyword = _sanitize_name(keyword)
    filepath = _get_output_path(f"cnki_titles_{clean_keyword}_{ts}.xlsx")
    wb = _build_single_sheet_workbook(results)

    saved_path = _try_save_workbook(wb, filepath, announce)
    save_result.record(saved_path)
    if saved_path and announce:
        _console.print("\n" + "═" * 50)
        _console.print(f"[bold green][*] 共抓取 {len(results)} 条数据。[/bold green]")
        _console.print(f"[*] 文件已保存至：")
        _console.print(f"    [bold]>>> {saved_path} <<<[/bold]")
        _console.print("═" * 50 + "\n")
    return save_result


def _save_multi_split(all_results: dict[str, list], ts: str, announce: bool):
    save_result = SaveResult()
    total = 0
    saved_files = []
    used_names: set[str] = set()
    for keyword, results in all_results.items():
        if not results:
            if announce:
                _console.print(f"[yellow][!] 关键词「{keyword}」无数据，跳过生成文件。[/yellow]")
            continue

        base_keyword = _sanitize_name(keyword)
        clean_keyword = base_keyword
        counter = 2
        while clean_keyword.casefold() in used_names:
            suffix = f"_{counter}"
            clean_keyword = f"{base_keyword[:50 - len(suffix)]}{suffix}"
            counter += 1
        if clean_keyword != base_keyword:
            _logger.warning(
                "分文件保存名冲突，已自动添加后缀: collision_index=%d",
                counter - 1,
            )
        used_names.add(clean_keyword.casefold())
        filepath = _get_output_path(f"cnki_titles_{clean_keyword}_{ts}.xlsx")
        wb = _build_single_sheet_workbook(results)
        saved_path = _try_save_workbook(wb, filepath, announce)
        save_result.record(saved_path)
        if saved_path:
            saved_files.append((keyword, len(results), saved_path))
            total += len(results)

    if announce:
        _console.print("\n" + "═" * 50)
        _console.print(
            f"[bold green][*] 全部抓取完毕，共 {total} 条数据，生成 {len(saved_files)} 个文件：[/bold green]"
        )
        for kw, cnt, path in saved_files:
            _console.print(f"  · [cyan][{kw}][/cyan] {cnt} 条  ->  {path}")
        _console.print("═" * 50 + "\n")
    return save_result


def _save_multi_merge(all_results: dict[str, list], ts: str, announce: bool):
    save_result = SaveResult()
    if not any(len(v) > 0 for v in all_results.values()):
        if announce:
            _console.print("[yellow][!] 所有关键词均未抓取到数据，不生成文件。[/yellow]")
        return save_result

    filepath = _get_output_path(f"cnki_titles_多词汇总_{ts}.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # noqa

    total = 0

    used_sheet_names: set[str] = set()
    for keyword, results in all_results.items():
        if not results:
            if announce:
                _console.print(f"[yellow][!] 关键词「{keyword}」无数据，跳过该 Sheet。[/yellow]")
            continue
        clean_keyword = _sanitize_name(keyword)
        base_name = clean_keyword[:31]
        sheet_name = base_name
        counter = 1
        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            sheet_name = base_name[:31 - len(suffix)] + suffix
            counter += 1
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append(_HEADERS)
        _append_records(ws, results)
        total += len(results)

    saved_path = _try_save_workbook(wb, filepath, announce)
    save_result.record(saved_path)
    if saved_path and announce:
        _console.print("\n" + "═" * 50)
        _console.print(f"[bold green][*] 全部抓取完毕，共 {total} 条数据。[/bold green]")
        _console.print(f"[*] 已合并保存至：")
        _console.print(f"    [bold]>>> {saved_path} <<<[/bold]")
        for kw, results in all_results.items():
            if not results:
                continue
            _console.print(f"  · Sheet [cyan][{kw}][/cyan]：{len(results)} 条")
        _console.print("═" * 50 + "\n")
    return save_result


def save_all(
    save_mode: str,
    keywords: list[str],
    all_results: dict[str, list],
    ts: str,
    announce: bool,
) -> SaveResult:
    """统一保存入口（幂等）。

    - 增量调用（announce=False）：静默写盘，用于每抓完一个关键词的阶段性落盘。
    - 最终调用（announce=True）：写盘并打印完整汇总，用于流程结束（含中断）时。

    文件名以传入的 ts 固定，增量与最终写同一文件、覆盖而非堆积，从而保证
    中途任何异常（含保存阶段的二次 Ctrl+C）都不会丢失已抓取的数据。

    claude opus4.8生成的神秘注释
    """
    if save_mode == "single":
        if keywords:
            return _save_single(keywords[0], all_results.get(keywords[0], []), ts, announce)
    elif save_mode == "multi_split":
        return _save_multi_split(all_results, ts, announce)
    elif save_mode == "multi_merge":
        return _save_multi_merge(all_results, ts, announce)
    return SaveResult()
